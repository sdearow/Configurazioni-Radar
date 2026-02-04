#!/usr/bin/env python3
"""
Generate a clean status entry table for intersection tracking.
This creates a simplified table for users to fill in status information.
"""

import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = '/home/user/Configurazioni-Radar/data-import'

def extract_code_and_name(postazione):
    if pd.isna(postazione) or not isinstance(postazione, str):
        return None, None
    match = re.match(r'^(\d+)\s*-\s*(.+)$', str(postazione).strip())
    if match:
        return match.group(1), match.group(2).strip()
    return None, str(postazione).strip()

def safe_str(val):
    if pd.isna(val):
        return ''
    return str(val).strip()

def main():
    print("Loading data sources...")

    # Load main LOTTI
    main_lotti_df = pd.read_excel(os.path.join(DATA_DIR, 'LOTTI M9_RADAR_v1_2026.01.28.xlsx'), sheet_name='INST.FIN. LOTTI')

    print("Creating status entry table...")
    status_data = []

    for _, row in main_lotti_df.iterrows():
        postazione = row.get('Numero-Nome Postazione')
        if pd.isna(postazione):
            continue

        code, name = extract_code_and_name(postazione)
        if not code or not name:
            continue

        lotto = row.get('Lotto')
        if lotto not in ['M9.1', 'M9.2']:
            continue

        record = {
            'CODE': code,
            'INTERSECTION_NAME': name,
            'LOTTO': safe_str(lotto),
            'SISTEMA': safe_str(row.get('Sistema')),
            'N_DISPOSITIVI': safe_str(row.get('N.ro Dispositivi')),

            # 4 Main Stage Status (for user to fill)
            'INSTALLATION': '',
            'CONFIGURATION': '',
            'CONNECTION': '',
            'VALIDATION': '',

            # Notes
            'NOTES': ''
        }

        status_data.append(record)

    # Create DataFrame
    df = pd.DataFrame(status_data)

    # Sort by Lotto, then by code
    def safe_sort_key(x):
        try:
            return int(float(x))
        except:
            return 9999

    df['sort_code'] = df['CODE'].apply(safe_sort_key)
    df = df.sort_values(['LOTTO', 'sort_code'])
    df = df.drop(columns=['sort_code'])

    # Save to Excel
    output_path = os.path.join(DATA_DIR, 'INTERSECTION_STATUS_TABLE.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # All intersections
        df.to_excel(writer, sheet_name='All Intersections', index=False)

        # M9.1 only
        m91 = df[df['LOTTO'] == 'M9.1']
        m91.to_excel(writer, sheet_name='M9.1', index=False)

        # M9.2 only
        m92 = df[df['LOTTO'] == 'M9.2']
        m92.to_excel(writer, sheet_name='M9.2', index=False)

    # Now format the Excel file
    wb = load_workbook(output_path)

    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    status_fill_pending = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_name in ['All Intersections', 'M9.1', 'M9.2']:
        ws = wb[sheet_name]

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Set column widths
        column_widths = {
            'A': 8,   # CODE
            'B': 40,  # INTERSECTION_NAME
            'C': 8,   # LOTTO
            'D': 10,  # SISTEMA
            'E': 12,  # N_DISPOSITIVI
            'F': 15,  # INSTALLATION
            'G': 15,  # CONFIGURATION
            'H': 15,  # CONNECTION
            'I': 15,  # VALIDATION
            'J': 30,  # NOTES
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Format data rows and status columns
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')

                # Highlight status columns with light yellow (pending to fill)
                if cell.column_letter in ['F', 'G', 'H', 'I']:
                    cell.fill = status_fill_pending

        # Freeze header row
        ws.freeze_panes = 'A2'

    wb.save(output_path)

    print(f"\nStatus table saved to: {output_path}")
    print(f"\nTotal intersections: {len(df)}")
    print(f"  - M9.1: {len(m91)}")
    print(f"  - M9.2: {len(m92)}")
    print(f"\nStatus columns to fill:")
    print("  - INSTALLATION: (e.g., Not Started, In Progress, Complete, Blocked)")
    print("  - CONFIGURATION: (e.g., Not Started, In Progress, Complete, Blocked)")
    print("  - CONNECTION: (e.g., Not Started, In Progress, Complete, Blocked)")
    print("  - VALIDATION: (e.g., Not Started, In Progress, Complete, Blocked)")

if __name__ == '__main__':
    main()
