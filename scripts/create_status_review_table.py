#!/usr/bin/env python3
"""
Create a review table for intersection status.
Shows interpreted status with underlying data for user verification.
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)

DATA_DIR = '/home/user/Configurazioni-Radar/data-import'

def safe_str(val):
    if pd.isna(val):
        return ''
    return str(val).strip()

def interpret_installation(row):
    """Interpret INSTALLATION status"""
    lotto = safe_str(row.get('LOTTO'))

    # Check for blocking issues first
    disp_bloccati = safe_str(row.get('DISP_INST_BLOCCATI'))

    if lotto == 'M9.1':
        l1_completato = safe_str(row.get('L1_COMPLETATO')).lower()
        l1_match = safe_str(row.get('L1_MATCH'))
        l1_install = safe_str(row.get('L1_INSTALL_SENSORI'))
        l1_cablaggio = safe_str(row.get('L1_CABLAGGIO'))

        if l1_completato == 'ok':
            return 'COMPLETE'
        elif l1_completato == 'parziale':
            return 'PARTIAL'
        elif l1_completato == 'no':
            return 'BLOCKED'
        elif disp_bloccati:
            return 'BLOCKED'
        elif l1_match:
            if l1_install and l1_cablaggio:
                return 'IN_PROGRESS'
            elif l1_install:
                return 'IN_PROGRESS'
            else:
                return 'STARTED'
        else:
            return 'NO_DATA'

    elif lotto == 'M9.2':
        l2_data = safe_str(row.get('L2_DATA_INSTALLAZ'))
        l2_radar = safe_str(row.get('L2_N_RADAR_FINITI'))
        l2_match = safe_str(row.get('L2_MATCH'))

        if l2_data and l2_radar:
            return 'COMPLETE'
        elif l2_data:
            return 'IN_PROGRESS'
        elif l2_match:
            return 'STARTED'
        else:
            return 'NO_DATA'

    return 'UNKNOWN'

def interpret_configuration(row):
    """Interpret CONFIGURATION status"""
    cfg_status = safe_str(row.get('CFG_DEF_STATUS')).lower()

    if cfg_status == 'ok':
        return 'COMPLETE'
    elif cfg_status == 'aff.go':
        return 'READY_FOR_APPROVAL'
    elif cfg_status == 'da vrf':
        return 'PENDING_VERIFICATION'
    elif cfg_status == 'inviata':
        return 'SENT'
    elif cfg_status == 'dainviare':
        return 'READY_TO_SEND'
    elif cfg_status == 'dafare':
        return 'NOT_STARTED'
    elif 'ass.' in cfg_status:
        return 'ASSIGNED'
    else:
        return 'UNKNOWN'

def interpret_connection(row):
    """Interpret CONNECTION status"""
    da_centr = safe_str(row.get('DA_CENTR_AUT')).lower()

    if 'centralizzato' in da_centr:
        return 'CENTRALIZED'
    elif da_centr == 'aut':
        return 'AUT_CONFIGURED'
    elif 'aut da install' in da_centr:
        return 'AUT_PENDING'
    elif da_centr == 'omnia':
        return 'OMNIA_PENDING'
    else:
        return 'NOT_STARTED'

def interpret_validation(row):
    """Interpret VALIDATION status"""
    vrf = safe_str(row.get('VRF_DATI'))

    if vrf and ('vrf' in vrf.lower() or 'utc' in vrf.lower()):
        return 'IN_VERIFICATION'
    else:
        return 'NOT_STARTED'

def main():
    print("Loading master intersection report...")
    df = pd.read_excel(os.path.join(DATA_DIR, 'MASTER_INTERSECTION_REPORT_CLEAN.xlsx'),
                       sheet_name='All Intersections')

    print(f"Creating review table for {len(df)} intersections...\n")

    results = []

    for idx, row in df.iterrows():
        code = safe_str(row.get('POSTAZIONE_CODE'))
        name = safe_str(row.get('POSTAZIONE_NAME'))
        lotto = safe_str(row.get('LOTTO'))
        sistema = safe_str(row.get('SISTEMA'))
        n_disp = safe_str(row.get('N_DISPOSITIVI'))

        # Get interpreted status
        inst_status = interpret_installation(row)
        cfg_status = interpret_configuration(row)
        conn_status = interpret_connection(row)
        val_status = interpret_validation(row)

        # Get key data for verification
        record = {
            'CODE': code,
            'NAME': name,
            'LOTTO': lotto,
            'SISTEMA': sistema,
            'N_DISP': n_disp,

            # INSTALLATION
            'INST_STATUS': inst_status,
            'INST_OK?': '',  # For user to confirm
            'INST_CORRECT': '',  # For user to provide correct value
        }

        # Add installation details based on lotto
        if lotto == 'M9.1':
            record['INST_DATA'] = f"L1: compl={safe_str(row.get('L1_COMPLETATO'))}, sens={safe_str(row.get('L1_INSTALL_SENSORI'))}, cabl={safe_str(row.get('L1_CABLAGGIO'))}, blocc={safe_str(row.get('DISP_INST_BLOCCATI'))}"
        else:
            record['INST_DATA'] = f"L2: data={safe_str(row.get('L2_DATA_INSTALLAZ'))}, radar={safe_str(row.get('L2_N_RADAR_FINITI'))}, centr={safe_str(row.get('L2_CENTRALIZZATI'))}"

        # CONFIGURATION
        record['CFG_STATUS'] = cfg_status
        record['CFG_OK?'] = ''
        record['CFG_CORRECT'] = ''
        record['CFG_DATA'] = f"status={safe_str(row.get('CFG_DEF_STATUS'))}, plan={safe_str(row.get('PLAN_CFG_INVIATE'))}, inst={safe_str(row.get('CFG_DEF_INST'))}"

        # CONNECTION
        record['CONN_STATUS'] = conn_status
        record['CONN_OK?'] = ''
        record['CONN_CORRECT'] = ''
        conn_data_parts = [
            f"centr={safe_str(row.get('DA_CENTR_AUT'))}",
            f"tab_utc={safe_str(row.get('TABELLA_IF_UTC'))}",
        ]
        if sistema.lower() == 'omnia':
            conn_data_parts.append(f"swarco={safe_str(row.get('SWARCO_SPOT_STATUS'))}")
        else:
            conn_data_parts.append(f"sema_aut={safe_str(row.get('SEMA_AUT'))}")
        record['CONN_DATA'] = ', '.join(conn_data_parts)

        # VALIDATION
        record['VAL_STATUS'] = val_status
        record['VAL_OK?'] = ''
        record['VAL_CORRECT'] = ''
        record['VAL_DATA'] = f"vrf={safe_str(row.get('VRF_DATI'))}"

        results.append(record)

    # Create DataFrame
    df_results = pd.DataFrame(results)

    # Save to Excel with formatting
    output_path = os.path.join(DATA_DIR, 'INTERSECTION_STATUS_REVIEW.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Status Review', index=False)

        # M9.1 only
        m91 = df_results[df_results['LOTTO'] == 'M9.1']
        m91.to_excel(writer, sheet_name='M9.1 Review', index=False)

        # M9.2 only
        m92 = df_results[df_results['LOTTO'] == 'M9.2']
        m92.to_excel(writer, sheet_name='M9.2 Review', index=False)

    # Format the Excel
    wb = load_workbook(output_path)

    # Style definitions
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    status_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    ok_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    data_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = border

        # Set column widths
        column_widths = {
            'A': 8,   # CODE
            'B': 35,  # NAME
            'C': 7,   # LOTTO
            'D': 8,   # SISTEMA
            'E': 7,   # N_DISP
            'F': 18,  # INST_STATUS
            'G': 8,   # INST_OK?
            'H': 15,  # INST_CORRECT
            'I': 60,  # INST_DATA
            'J': 22,  # CFG_STATUS
            'K': 8,   # CFG_OK?
            'L': 15,  # CFG_CORRECT
            'M': 50,  # CFG_DATA
            'N': 18,  # CONN_STATUS
            'O': 8,   # CONN_OK?
            'P': 15,  # CONN_CORRECT
            'Q': 60,  # CONN_DATA
            'R': 15,  # VAL_STATUS
            'S': 8,   # VAL_OK?
            'T': 15,  # VAL_CORRECT
            'U': 20,  # VAL_DATA
        }

        for col_letter, width in column_widths.items():
            if col_letter in 'ABCDEFGHIJKLMNOPQRSTU':
                ws.column_dimensions[col_letter].width = width

        # Format data rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                col_letter = get_column_letter(col_idx)

                # Status columns (green)
                if col_letter in ['F', 'J', 'N', 'R']:
                    cell.fill = status_fill
                # OK? columns (yellow - for user input)
                elif col_letter in ['G', 'K', 'O', 'S']:
                    cell.fill = ok_fill
                # CORRECT columns (yellow - for user input)
                elif col_letter in ['H', 'L', 'P', 'T']:
                    cell.fill = ok_fill
                # DATA columns (gray)
                elif col_letter in ['I', 'M', 'Q', 'U']:
                    cell.fill = data_fill
                    cell.font = Font(size=9)

        # Freeze header and first columns
        ws.freeze_panes = 'C2'

        # Set row height
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 30

    wb.save(output_path)

    print(f"Review table saved to: {output_path}")
    print(f"\nColumns explanation:")
    print("  - *_STATUS: My interpreted status based on data")
    print("  - *_OK?: Fill with Y if correct, N if incorrect")
    print("  - *_CORRECT: If not OK, fill with correct status")
    print("  - *_DATA: Raw data used for interpretation")

    # Print summary
    print("\n" + "=" * 80)
    print("SUMMARY OF INTERPRETED STATUS")
    print("=" * 80)

    print("\nINSTALLATION:")
    print(df_results['INST_STATUS'].value_counts().to_string())

    print("\nCONFIGURATION:")
    print(df_results['CFG_STATUS'].value_counts().to_string())

    print("\nCONNECTION:")
    print(df_results['CONN_STATUS'].value_counts().to_string())

    print("\nVALIDATION:")
    print(df_results['VAL_STATUS'].value_counts().to_string())

if __name__ == '__main__':
    main()
