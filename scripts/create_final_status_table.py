#!/usr/bin/env python3
"""
Create a review table for intersection status with corrected CONNECTION logic.
Uses SWARCO file for Omnia and Semaforica file for Tmacs connection status.
"""

import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)

DATA_DIR = '/home/user/Configurazioni-Radar/data-import'

def safe_str(val):
    if pd.isna(val):
        return ''
    return str(val).strip()

def extract_code(postazione):
    if pd.isna(postazione) or not isinstance(postazione, str):
        return None
    match = re.match(r'^(\d+)', str(postazione).strip())
    if match:
        return match.group(1)
    return None

def load_swarco_data():
    """Load SWARCO data and create lookup by codice impianto"""
    swarco = pd.read_excel(os.path.join(DATA_DIR, 'Swarco_Verifica stato - TOT_2026.01.29.xlsx'),
                           sheet_name='Foglio1')

    # The key column for status
    spot_col = 'SPOT:\nSu scheda (centralino SCAE)?\nCon firmware vecchio (da sostituire)?\nCon firmware recente (aggiornabile da remoto)?'

    lookup = {}
    for _, row in swarco.iterrows():
        codice = row.get('Codice\nImpianto')
        if pd.isna(codice):
            continue
        codice_str = str(int(float(codice)))

        spot_status = safe_str(row.get(spot_col))
        lookup[codice_str] = spot_status

    return lookup

def load_semaforica_data():
    """Load Semaforica data and create lookup by codice impianto"""
    sema = pd.read_excel(os.path.join(DATA_DIR, 'Elenco IS RADAR_SEMAFORICA_per VERIFICA STATO AUT.xlsx'),
                         sheet_name='IMPIANTI SEMAFORICA')

    lookup = {}
    for _, row in sema.iterrows():
        codice = row.get('Codice\nImpianto')
        if pd.isna(codice):
            continue
        # Normalize: remove leading zeros by converting to int then back to string
        try:
            codice_str = str(int(codice))
        except:
            codice_str = str(codice).strip().lstrip('0') or '0'

        aut = safe_str(row.get('AUT'))
        attivita = safe_str(row.get("ATTIVITA' DA FARE"))
        lookup[codice_str] = {'AUT': aut, 'ATTIVITA': attivita}

    return lookup

def interpret_installation(row):
    """Interpret INSTALLATION status"""
    lotto = safe_str(row.get('LOTTO'))
    disp_bloccati = safe_str(row.get('DISP_INST_BLOCCATI'))

    if lotto == 'M9.1':
        l1_completato = safe_str(row.get('L1_COMPLETATO')).lower()
        l1_match = safe_str(row.get('L1_MATCH'))
        l1_install = safe_str(row.get('L1_INSTALL_SENSORI'))
        l1_cablaggio = safe_str(row.get('L1_CABLAGGIO'))

        if l1_completato == 'ok':
            return 'COMPLETE'
        elif l1_completato == 'parziale':
            return 'PARTIAL'
        elif l1_completato == 'no':
            return 'BLOCKED'
        elif disp_bloccati:
            return 'BLOCKED'
        elif l1_match:
            if l1_install and l1_cablaggio:
                return 'IN_PROGRESS'
            elif l1_install:
                return 'IN_PROGRESS'
            else:
                return 'STARTED'
        else:
            return 'NO_DATA'

    elif lotto == 'M9.2':
        l2_data = safe_str(row.get('L2_DATA_INSTALLAZ'))
        l2_radar = safe_str(row.get('L2_N_RADAR_FINITI'))
        l2_match = safe_str(row.get('L2_MATCH'))

        if l2_data and l2_radar:
            return 'COMPLETE'
        elif l2_data:
            return 'IN_PROGRESS'
        elif l2_match:
            return 'STARTED'
        else:
            return 'NO_DATA'

    return 'UNKNOWN'

def interpret_configuration(row):
    """Interpret CONFIGURATION status"""
    cfg_status = safe_str(row.get('CFG_DEF_STATUS')).lower()

    if cfg_status == 'ok':
        return 'COMPLETE'
    elif cfg_status == 'aff.go':
        return 'READY_FOR_APPROVAL'
    elif cfg_status == 'da vrf':
        return 'PENDING_VERIFICATION'
    elif cfg_status == 'inviata':
        return 'SENT'
    elif cfg_status == 'dainviare':
        return 'READY_TO_SEND'
    elif cfg_status == 'dafare':
        return 'NOT_STARTED'
    elif 'ass.' in cfg_status:
        return 'ASSIGNED'
    else:
        return 'UNKNOWN'

def interpret_connection_omnia(codice_impianto, swarco_lookup):
    """
    Interpret CONNECTION status for Omnia intersections using SWARCO data.

    Rules:
    - Idoneo → IN_PROGRESS
    - Non Idoneo - SHDSL da aggiornare → BLOCKED
    - Impossibile connettersi con SSH → BLOCKED
    - Impianto aggiunto ma SPOT non presente → BLOCKED
    - Not in file → MISSING_DATA
    """
    codice_str = str(int(float(codice_impianto))) if codice_impianto else ''

    if codice_str not in swarco_lookup:
        return 'MISSING_DATA', ''

    spot_status = swarco_lookup[codice_str]

    if not spot_status:
        return 'MISSING_DATA', spot_status

    spot_lower = spot_status.lower()

    if spot_lower == 'idoneo':
        return 'IN_PROGRESS', spot_status
    elif 'non idoneo' in spot_lower or 'shdsl' in spot_lower:
        return 'BLOCKED', spot_status
    elif 'impossibile' in spot_lower or 'ssh' in spot_lower:
        return 'BLOCKED', spot_status
    elif 'spot' in spot_lower and ('non presente' in spot_lower or 'sim' in spot_lower):
        return 'BLOCKED', spot_status
    else:
        return 'UNKNOWN', spot_status

def interpret_connection_tmacs(codice_impianto, sema_lookup):
    """
    Interpret CONNECTION status for Tmacs intersections using Semaforica data.

    Rules:
    - AUT = NO → BLOCKED
    - AUT = SI with "Configurare AUT da remoto/in lab/Sostituire AUT?!?!" → NO_INFO
    - AUT = SI with "AUT da aggiornare" → IN_PROGRESS
    - AUT = SI with "AUT da SOSTITUIRE" → BLOCKED
    - Not in file → MISSING_DATA
    """
    # Normalize codice: convert to int to remove decimals and leading zeros
    try:
        codice_str = str(int(float(codice_impianto)))
    except:
        codice_str = ''

    if codice_str not in sema_lookup:
        return 'MISSING_DATA', ''

    data = sema_lookup[codice_str]
    aut = data['AUT'].upper()
    attivita = data['ATTIVITA']

    info_str = f"AUT={aut}, ATTIVITA={attivita}"

    if aut == 'NO':
        return 'BLOCKED', info_str
    elif aut == 'SI':
        attivita_lower = attivita.lower()
        if 'configurare aut da remoto' in attivita_lower or 'configurare aut in lab' in attivita_lower or 'sostituire aut?!?!' in attivita_lower:
            return 'NO_INFO', info_str
        elif 'aut da sostituire' in attivita_lower or 'troppo vecchia' in attivita_lower:
            return 'BLOCKED', info_str
        elif 'aut da aggiornare' in attivita_lower:
            return 'IN_PROGRESS', info_str
        elif 'con aut' in attivita_lower:
            return 'IN_PROGRESS', info_str
        elif 'senza aut' in attivita_lower:
            return 'BLOCKED', info_str
        else:
            return 'UNKNOWN', info_str
    else:
        return 'UNKNOWN', info_str

def interpret_validation(row):
    """Interpret VALIDATION status"""
    vrf = safe_str(row.get('VRF_DATI'))

    if vrf and ('vrf' in vrf.lower() or 'utc' in vrf.lower()):
        return 'IN_VERIFICATION'
    else:
        return 'NOT_STARTED'

def main():
    print("Loading data sources...")

    # Load main data
    df = pd.read_excel(os.path.join(DATA_DIR, 'MASTER_INTERSECTION_REPORT_CLEAN.xlsx'),
                       sheet_name='All Intersections')

    # Load connection status lookups
    swarco_lookup = load_swarco_data()
    sema_lookup = load_semaforica_data()

    print(f"SWARCO lookup: {len(swarco_lookup)} entries")
    print(f"Semaforica lookup: {len(sema_lookup)} entries")
    print(f"\nAnalyzing {len(df)} intersections...\n")

    results = []

    for idx, row in df.iterrows():
        code = safe_str(row.get('POSTAZIONE_CODE'))
        name = safe_str(row.get('POSTAZIONE_NAME'))
        lotto = safe_str(row.get('LOTTO'))
        sistema = safe_str(row.get('SISTEMA')).lower()
        n_disp = safe_str(row.get('N_DISPOSITIVI'))
        codice_impianto = row.get('CODICE_IMPIANTO')

        # Get interpreted status
        inst_status = interpret_installation(row)
        cfg_status = interpret_configuration(row)
        val_status = interpret_validation(row)

        # CONNECTION status based on sistema
        if sistema == 'omnia':
            conn_status, conn_detail = interpret_connection_omnia(codice_impianto, swarco_lookup)
        elif sistema == 'tmacs':
            conn_status, conn_detail = interpret_connection_tmacs(codice_impianto, sema_lookup)
        else:
            conn_status, conn_detail = 'UNKNOWN', f'Sistema: {sistema}'

        # Build installation data string
        if lotto == 'M9.1':
            inst_data = f"compl={safe_str(row.get('L1_COMPLETATO'))}, sens={safe_str(row.get('L1_INSTALL_SENSORI'))}, cabl={safe_str(row.get('L1_CABLAGGIO'))}, blocc={safe_str(row.get('DISP_INST_BLOCCATI'))}"
        else:
            inst_data = f"data={safe_str(row.get('L2_DATA_INSTALLAZ'))}, radar={safe_str(row.get('L2_N_RADAR_FINITI'))}, centr={safe_str(row.get('L2_CENTRALIZZATI'))}"

        record = {
            'CODE': code,
            'NAME': name,
            'LOTTO': lotto,
            'SISTEMA': sistema.upper(),
            'N_DISP': n_disp,
            'CODICE_IMPIANTO': str(int(float(codice_impianto))) if codice_impianto and not pd.isna(codice_impianto) else '',

            # INSTALLATION
            'INST_STATUS': inst_status,
            'INST_DATA': inst_data,

            # CONFIGURATION
            'CFG_STATUS': cfg_status,
            'CFG_DATA': f"status={safe_str(row.get('CFG_DEF_STATUS'))}, plan={safe_str(row.get('PLAN_CFG_INVIATE'))}",

            # CONNECTION
            'CONN_STATUS': conn_status,
            'CONN_DATA': conn_detail,

            # VALIDATION
            'VAL_STATUS': val_status,
            'VAL_DATA': f"vrf={safe_str(row.get('VRF_DATI'))}",
        }

        results.append(record)

    # Create DataFrame
    df_results = pd.DataFrame(results)

    # Save to Excel with formatting
    output_path = os.path.join(DATA_DIR, 'INTERSECTION_STATUS_FINAL.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='All Intersections', index=False)

        # M9.1 only
        m91 = df_results[df_results['LOTTO'] == 'M9.1']
        m91.to_excel(writer, sheet_name='M9.1', index=False)

        # M9.2 only
        m92 = df_results[df_results['LOTTO'] == 'M9.2']
        m92.to_excel(writer, sheet_name='M9.2', index=False)

    # Format the Excel
    wb = load_workbook(output_path)

    # Style definitions
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)

    # Status colors
    complete_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
    progress_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow
    blocked_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')   # Red
    nodata_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')    # Gray

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    status_colors = {
        'COMPLETE': complete_fill,
        'IN_PROGRESS': progress_fill,
        'IN_VERIFICATION': progress_fill,
        'READY_FOR_APPROVAL': progress_fill,
        'PENDING_VERIFICATION': progress_fill,
        'SENT': progress_fill,
        'ASSIGNED': progress_fill,
        'READY_TO_SEND': progress_fill,
        'STARTED': progress_fill,
        'PARTIAL': progress_fill,
        'BLOCKED': blocked_fill,
        'NOT_STARTED': nodata_fill,
        'NO_DATA': nodata_fill,
        'MISSING_DATA': nodata_fill,
        'NO_INFO': nodata_fill,
        'UNKNOWN': nodata_fill,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = border

        # Set column widths
        column_widths = {
            'A': 8,   # CODE
            'B': 35,  # NAME
            'C': 7,   # LOTTO
            'D': 8,   # SISTEMA
            'E': 7,   # N_DISP
            'F': 15,  # CODICE_IMPIANTO
            'G': 20,  # INST_STATUS
            'H': 50,  # INST_DATA
            'I': 22,  # CFG_STATUS
            'J': 40,  # CFG_DATA
            'K': 15,  # CONN_STATUS
            'L': 60,  # CONN_DATA
            'M': 15,  # VAL_STATUS
            'N': 20,  # VAL_DATA
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Format data rows with color coding
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                col_letter = get_column_letter(col_idx)

                # Color code status columns
                if col_letter in ['G', 'I', 'K', 'M']:  # Status columns
                    status_val = cell.value
                    if status_val in status_colors:
                        cell.fill = status_colors[status_val]
                        cell.font = Font(bold=True)

        # Freeze header and first columns
        ws.freeze_panes = 'C2'

    wb.save(output_path)

    print(f"Final status table saved to: {output_path}")

    # Print summary
    print("\n" + "=" * 80)
    print("SUMMARY OF STATUS")
    print("=" * 80)

    print("\nINSTALLATION:")
    print(df_results['INST_STATUS'].value_counts().to_string())

    print("\nCONFIGURATION:")
    print(df_results['CFG_STATUS'].value_counts().to_string())

    print("\nCONNECTION:")
    print(df_results['CONN_STATUS'].value_counts().to_string())

    print("\nVALIDATION:")
    print(df_results['VAL_STATUS'].value_counts().to_string())

    # Print detailed connection breakdown by sistema
    print("\n" + "=" * 80)
    print("CONNECTION STATUS BY SISTEMA")
    print("=" * 80)

    print("\nOMNIA:")
    omnia = df_results[df_results['SISTEMA'] == 'OMNIA']
    print(omnia['CONN_STATUS'].value_counts().to_string())

    print("\nTMACS:")
    tmacs = df_results[df_results['SISTEMA'] == 'TMACS']
    print(tmacs['CONN_STATUS'].value_counts().to_string())

if __name__ == '__main__':
    main()
